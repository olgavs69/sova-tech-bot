from asyncio import get_event_loop
from aiogram import Router, F, types
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, CallbackQuery, BufferedInputFile, FSInputFile
from aiogram.types import InlineKeyboardMarkup as IKM, InlineKeyboardButton as IKB

import pandas as pd
from io import BytesIO
import json
import os
import locale

from reportlab import pdfbase
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab import pdfbase
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import TableStyle, Table
from reportlab.lib import colors

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from src.basic.graphics.graphicsforpdf import create_combined_graph
from src.mailing.data.notification.notification_google_sheets_worker import notification_gsworker
from src.mailing.data.techsupport.techsupport_google_sheets_worker import techsupport_gsworker
from src.analytics.db.db import user_tokens_db
from src.util.log import logger

from src.basic.keyboards.keyboards import get_markup, get_report_format_keyboard

router = Router(name=__name__)

pdfmetrics.registerFont(TTFont('FreeSerif', 'resources/fonts/FreeSerif.ttf'))  # Убедитесь, что путь к шрифту правильный
# canvas.setFont

@router.callback_query(F.data == 'start')
async def start_callback_handler(query: CallbackQuery, state: FSMContext) -> None:
    await start_handler(query.from_user.id, query.message, state)
    await query.answer()


@router.message(CommandStart())
async def command_start_handler(message: types.Message, state: FSMContext) -> None:
    await start_handler(message.from_user.id, message, state)


async def start_handler(user_id: int, message: types.Message, state: FSMContext) -> None:
    await state.clear()

    msg = await message.answer("Загрузка... ⚙️")

    loop = get_event_loop()
    has_token = user_tokens_db.has_tgid(user_id)
    kb = await loop.run_in_executor(None, get_markup, user_id, has_token)

    await msg.edit_text(
        text=f"Вас приветствует чат-бот SOVA-tech!",
        reply_markup=kb,
    )


@router.callback_query(F.data == 'generate_report')
async def generate_report_handler(query: CallbackQuery) -> None:
    print("Generate report callback triggered")  # Debug line
    btn_pdf = [IKB(text='PDF 📄', callback_data='generate_report_pdf')]
    btn_excel = [IKB(text='Excel 📊', callback_data='generate_report_excel')]
    kb = IKM(inline_keyboard=[btn_pdf, btn_excel])

    # Обновляем сообщение с новыми кнопками
    await query.message.edit_text(
        text="Выберите формат отчёта:",
        reply_markup=kb
    )
    await query.answer()



def create_report_pdf(data: dict) -> BytesIO:
    file_bytes = BytesIO()
    try:
        c = canvas.Canvas(file_bytes, pagesize=letter)
        width, height = letter

        # Use built-in fonts (e.g., Helvetica)
        c.setFont("Helvetica", 12)  # Use a built-in font

        # Заголовок отчёта
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, height - 50, "ОТЧЁТ ВЫРУЧКА (анализ)")

        y_position = height - 100

        # Проходим по данным и выводим их на странице
        for section_title, section_data in data.items():
            c.setFont("Helvetica-Bold", 12)
            c.drawString(100, y_position, section_title)
            y_position -= 20

            for item in section_data:
                c.setFont("Helvetica", 10)
                c.drawString(100, y_position, item)
                y_position -= 15

            # Добавляем отступ между разделами
            y_position -= 10

        c.showPage()
        c.save()

        file_bytes.seek(0)
    except Exception as e:
        raise Exception(f"Ошибка при создании PDF: {e}")

    return file_bytes


def create_empty_excel() -> BytesIO:
    file_bytes = BytesIO()
    try:
        df = pd.DataFrame({
            "Дата": ["01.01.2025"],
            "Заголовок отчёта": ["Отчёт по аналитике"],
            "Описание": ["Пустой отчёт с шаблоном данных"]
        })

        with pd.ExcelWriter(file_bytes, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Отчёт')

        file_bytes.seek(0)
    except Exception as e:
        raise Exception(f"Ошибка при создании Excel: {e}")

    return file_bytes

def create_empty_pdf() -> BytesIO:
    file_bytes = BytesIO()
    try:
        c = canvas.Canvas(file_bytes, pagesize=letter)  # Размер страницы Letter
        c.setFont("Helvetica", 12)  # Используем стандартный шрифт Helvetica
        c.drawString(100, 750, "Этот файл был сгенерирован ботом.")
        c.drawString(100, 730, "Пустой отчёт PDF с шаблоном данных.")
        c.showPage()
        c.save()

        file_bytes.seek(0)  # Сдвиг указателя в начало
    except Exception as e:
        raise Exception(f"Ошибка при создании PDF: {e}")

    return file_bytes


@router.callback_query(F.data == 'generate_report_pdf')
async def send_empty_pdf_report(query: CallbackQuery) -> None:
    try:
        pdf_file = create_empty_pdf()

        input_file = BufferedInputFile(pdf_file.getvalue(), filename="empty_report.pdf")

        await query.message.answer_document(
            document=input_file,
            caption="Ваш пустой отчёт в формате PDF готов!"
        )
    except Exception as e:
        await query.message.answer(f"Произошла ошибка при создании PDF отчёта: {e}")



@router.callback_query(F.data == 'generate_report_excel')
async def send_empty_excel_report(query: CallbackQuery) -> None:
    try:
        excel_file = create_empty_excel()

        input_file = BufferedInputFile(excel_file.getvalue(), filename="empty_report.xlsx")

        await query.message.answer_document(
            document=input_file,
            caption="Ваш пустой отчёт в формате Excel готов!"
        )
    except Exception as e:
        await query.message.answer(f"Произошла ошибка при создании Excel отчёта: {e}")


# Выгрузка отчёта Анализ выручки

def create_json_report_pdf(data: dict) -> BytesIO:
    file_bytes = BytesIO()
    try:
        c = canvas.Canvas(file_bytes, pagesize=letter)
        width, height = letter
        margin = 50  # Page margins
        y_position = height - 50  # Initial text position

        # Title of the report
        c.setFont("FreeSerif", 16)
        c.drawString(margin, y_position, "АНАЛИЗ ВЫРУЧКИ")
        y_position -= 60  # Increased space after the title

        c.setFont("FreeSerif", 7)  # Reduced font size for the table
        max_text_width = width - 2 * margin  # Maximum text width

        # Processing store data and displaying the table
        data_table = []
        headers = ["Магазин", "Выручка", "Динамика (неделя)", "Динамика (месяц)", "Динамика (год)", "Прогноз"]
        data_table.append(headers)

        # Add rows with formatted numbers
        for store in data["data"]:
            row = [
                store["label"],
                format_number(store["revenue"]),
                format_number(store["revenue_dynamics_week"]),
                format_number(store["revenue_dynamics_month"]),
                format_number(store["revenue_dynamics_year"]),
                format_number(store["revenue_forecast"])
            ]
            data_table.append(row)

        # Add totals with formatted numbers
        totals = [
            data["sum"]["label"],
            format_number(data["sum"]["revenue"]),
            format_number(data["sum"]["revenue_dynamics_week"]),
            format_number(data["sum"]["revenue_dynamics_month"]),
            format_number(data["sum"]["revenue_dynamics_year"]),
            format_number(data["sum"]["revenue_forecast"])
        ]
        data_table.append(totals)

        # Create the table with the collected data and adjusted column widths
        table = Table(data_table, colWidths=[140, 80, 80, 80, 80, 80])  # Adjusted column widths
        table.setStyle(TableStyle([  # Styling the table
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), "FreeSerif"),
            ('FONTNAME', (0, 1), (-1, -1), "FreeSerif"),
            ('FONTSIZE', (0, 0), (-1, -1), 7),  # Reduced font size for table text
        ]))

        # Add table to PDF
        table.wrapOn(c, width, height)
        table.drawOn(c, margin, y_position - len(data_table) * 12)  # Adjust row height to 12

        # Update y_position after the table
        y_position -= len(data_table) * 12 + 50  # Space after table before graphs

        # Create and add the combined graph (bar + pie chart)
        combined_graph_bytes = create_combined_graph(data)  # Call the function to generate the combined graph
        combined_graph_img = ImageReader(combined_graph_bytes)

        # Adjust width and position for a wider graph
        graph_width = width - 100  # Wider graph
        graph_height = 400  # Adjusted graph height
        graph_y_position = y_position - graph_height  # Adjusted y position to place the graph below the table

        # Draw the image (graph) on the canvas
        c.drawImage(combined_graph_img, 50, graph_y_position, graph_width, graph_height)  # Wider graph

        c.showPage()  # End the page
        c.save()

        file_bytes.seek(0)

    except Exception as e:
        raise Exception(f"Ошибка при создании PDF: {e}")

    return file_bytes


# Пример данных для отчёта
report_data = {
    "1 Гостепоток и средний чек": [
        "- средний чек: -10%, 500/460 руб",
        "глубина чека: -4%, 4,5/3,8 позиций в чеке",
        "+ количество чеков +7%, 1200/1600 чеков"
    ],
    "2 Выручка по направлениям": [
        "- бар -9%, 38%/33% доля бара",
        "+ кухня +3%, 62%/67% доля кухни"
    ],
    "3 Выручка по группам блюд": [
        "- салаты -7%, 450 000/410 000 руб",
        "супы -12%, 250 000/190 000 руб",
        "+ выпечка +10%, 500 000/600 000 руб",
        "кофе +15%, 400 000/550 000 руб"
    ],
    # Добавьте остальные секции отчёта аналогично
}

def create_report_excel(data: dict) -> BytesIO:
    file_bytes = BytesIO()
    try:
        with pd.ExcelWriter(file_bytes, engine='xlsxwriter') as writer:
            for section_title, section_data in data.items():
                # Создаём DataFrame для каждой секции отчёта
                df = pd.DataFrame(section_data, columns=[section_title])

                # Пишем данные в новый лист Excel
                df.to_excel(writer, index=False, sheet_name=section_title)

        file_bytes.seek(0)
    except Exception as e:
        raise Exception(f"Ошибка при создании Excel: {e}")

    return file_bytes

# Пример данных для отчёта в формате Excel
report_data_excel = {
    "1 Гостепоток и средний чек": [
        ["- средний чек: -10%, 500/460 руб"],
        ["глубина чека: -4%, 4,5/3,8 позиций в чеке"],
        ["+ количество чеков +7%, 1200/1600 чеков"]
    ],
    "2 Выручка по направлениям": [
        ["- бар -9%, 38%/33% доля бара"],
        ["+ кухня +3%, 62%/67% доля кухни"]
    ],
    "3 Выручка по группам блюд": [
        ["- салаты -7%, 450 000/410 000 руб"],
        ["супы -12%, 250 000/190 000 руб"],
        ["+ выпечка +10%, 500 000/600 000 руб"],
        ["кофе +15%, 400 000/550 000 руб"]
    ],
    # Добавьте остальные секции отчёта аналогично
}

# Генерация отчёта в Excel
excel_file = create_report_excel(report_data_excel)



@router.callback_query(F.data == 'generate_sample_report')
async def generate_sample_report_handler(query: CallbackQuery) -> None:
    # Логика для отображения кнопок выбора формата отчёта
    btn_pdf = [IKB(text='PDF 📄', callback_data='generate_sample_report_pdf')]
    btn_excel = [IKB(text='Excel 📊', callback_data='generate_sample_report_excel')]
    kb = IKM(inline_keyboard=[btn_pdf, btn_excel])

    # Обновляем сообщение с новыми кнопками
    await query.message.edit_text(
        text="Выберите формат примерного отчёта:",
        reply_markup=kb
    )
    await query.answer()

@router.callback_query(F.data == 'generate_sample_report_pdf')
async def generate_sample_report_pdf_handler(query: CallbackQuery) -> None:
    try:
        # Генерация примерного отчёта в PDF
        pdf_file = create_report_pdf(report_data)

        input_file = BufferedInputFile(pdf_file.getvalue(), filename="sample_report_выручка.pdf")

        await query.message.answer_document(
            document=input_file,
            caption="Примерный отчёт по выручке (анализ) в формате PDF готов!"
        )
    except Exception as e:
        await query.message.answer(f"Произошла ошибка при создании примерного PDF отчёта: {e}")


@router.callback_query(F.data == 'generate_sample_report_excel')
async def generate_sample_report_excel_handler(query: CallbackQuery) -> None:
    try:
        # Генерация примерного отчёта в Excel
        excel_file = create_report_excel(report_data_excel)

        input_file = BufferedInputFile(excel_file.getvalue(), filename="sample_report_выручка.xlsx")

        await query.message.answer_document(
            document=input_file,
            caption="Примерный отчёт по выручке (анализ) в формате Excel готов!"
        )
    except Exception as e:
        await query.message.answer(f"Произошла ошибка при создании примерного Excel отчёта: {e}")


JSON_FILE_PATH = "files/example.json"

def load_json_from_file():
    """Загружает данные из example.json, если он есть"""
    if os.path.exists(JSON_FILE_PATH):
        try:
            with open(JSON_FILE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError:
            print("Ошибка при чтении JSON-файла.")
            return {}
    print("Файл JSON не найден.")
    return {}


@router.message(F.text == 'Загрузить данные')
async def load_data_handler(message: Message, state: FSMContext):
    # Загрузка и обработка данных (например, JSON)
    # Предполагаем, что данные уже загружены в состояние
    await state.update_data(json_data={"key": "value"})  # Пример данных

    # Отправляем пользователю кнопки для выбора формата
    await message.answer(
        "Данные загружены. Выберите формат для выгрузки отчёта:",
        reply_markup=get_report_format_keyboard()
    )

@router.callback_query(F.data == 'generate_json_report')
async def generate_json_report_handler(query: CallbackQuery, state: FSMContext):
    # Загружаем данные из example.json
    json_data = load_json_from_file()

    if not json_data:
        await query.message.answer("Данные не найдены в JSON файле.")
        return

    # Сохраняем данные в FSMContext
    await state.update_data(json_data=json_data)

    # Клавиатура для выбора формата отчёта
    kb = IKM(inline_keyboard=[[
        IKB(text='PDF 📄', callback_data='generate_json_report_pdf')
    ], [
        IKB(text='Excel 📊', callback_data='generate_json_report_excel')
    ]])

    # Отправляем пользователю сообщение с кнопками для выбора формата
    await query.message.edit_text("Выберите формат отчёта:", reply_markup=kb)
    await query.answer()


@router.message(F.data == 'generate_json_report')
async def process_json_file(message: Message, state: FSMContext):
    document = message.document
    if not document.file_name.endswith('.json'):
        await message.answer("Пожалуйста, отправьте файл в формате JSON.")
        return

    file = await message.bot.get_file(document.file_id)
    file_path = file.file_path
    file_bytes = await message.bot.download_file(file_path)

    try:
        uploaded_json = json.load(BytesIO(file_bytes))
    except json.JSONDecodeError:
        await message.answer("Ошибка при чтении JSON-файла. Убедитесь, что формат корректный.")
        return

    # Загружаем данные из example.json
    file_json = load_json_from_file()

    # Объединяем данные
    combined_data = {**file_json, **uploaded_json}

    # Сохраняем данные в FSMContext
    await state.update_data(json_data=combined_data)

    # Выводим отладочную информацию
    print("Объединенные данные:", combined_data)

    # Клавиатура для выбора формата отчёта
    kb = IKM(inline_keyboard=[
        [IKB(text='PDF 📄', callback_data='generate_json_report_pdf')],
        [IKB(text='Excel 📊', callback_data='generate_json_report_excel')]
    ])
    await message.answer("Выберите формат отчёта:", reply_markup=kb)


def format_number(value):
    """Helper function to format numbers"""
    return f"{value:,.2f}"


def create_report_excel(data: dict) -> BytesIO:
    # Create a new workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Define the headers
    headers = ["Магазин", "Выручка", "Динамика (неделя)", "Динамика (месяц)", "Динамика (год)", "Прогноз"]

    # Add headers to the first row and format them
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)  # Bold headers
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

    # Adding store data rows
    row_num = 2  # Start from the second row (after the header)
    for store in data["data"]:
        row = [
            store["label"],
            format_number(store["revenue"]),
            format_number(store["revenue_dynamics_week"]),
            format_number(store["revenue_dynamics_month"]),
            format_number(store["revenue_dynamics_year"]),
            format_number(store["revenue_forecast"])
        ]
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment
        row_num += 1

    # Add the totals row
    totals = [
        data["sum"]["label"],
        format_number(data["sum"]["revenue"]),
        format_number(data["sum"]["revenue_dynamics_week"]),
        format_number(data["sum"]["revenue_dynamics_month"]),
        format_number(data["sum"]["revenue_dynamics_year"]),
        format_number(data["sum"]["revenue_forecast"])
    ]
    for col_num, value in enumerate(totals, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Center alignment

    # Save the workbook to a BytesIO object to send as a document
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  # Reset the pointer to the beginning of the file
    return excel_file


@router.callback_query(F.data == 'generate_json_report_excel')
async def generate_json_report_excel_handler(query: CallbackQuery, state: FSMContext):
    try:
        # Получаем данные из состояния
        user_data = await state.get_data()
        json_data = user_data.get('json_data', {})

        if not json_data:
            await query.message.answer("Нет данных для генерации отчёта. Пожалуйста, загрузите данные.")
            return

        # Генерация Excel отчёта
        excel_file = create_report_excel(json_data)
        input_file = BufferedInputFile(excel_file.getvalue(), filename="анализ_выручки.xlsx")  # Название файла

        # Отправляем Excel отчёт
        await query.message.answer_document(
            document=input_file,
            caption="Отчёт в формате Excel готов!"
        )
    except Exception as e:
        await query.message.answer(f"Ошибка при создании Excel отчёта: {e}")


locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')

def format_number(number):
    # Format the number with thousand separators (dot as thousands separator)
    try:
        return locale.format_string("%d", number, grouping=True)
    except Exception:
        return str(number)

@router.callback_query(F.data == 'generate_json_report_pdf')
async def generate_json_report_pdf_handler(query: CallbackQuery, state: FSMContext):
    try:
        # Получаем данные из состояния
        user_data = await state.get_data()
        json_data = user_data.get('json_data', {})

        if not json_data:
            await query.message.answer("Нет данных для генерации отчёта. Пожалуйста, загрузите данные.")
            return

        # Генерация PDF отчёта
        pdf_file = create_json_report_pdf(json_data)
        input_file = BufferedInputFile(pdf_file.getvalue(), filename="json_report.pdf")

        # Отправляем PDF отчёт
        await query.message.answer_document(
            document=input_file,
            caption="Отчёт по данным JSON в формате PDF готов!"
        )
    except Exception as e:
        await query.message.answer(f"Ошибка при создании PDF отчёта: {e}")


def wrap_text(text, font, font_size, max_width):
    """Функция для переноса текста, если он слишком длинный для страницы"""
    lines = []
    current_line = []
    current_width = 0

    for word in text.split():
        word_width = pdfmetrics.stringWidth(word, font, font_size)
        if current_width + word_width <= max_width:
            current_line.append(word)
            current_width += word_width + pdfmetrics.stringWidth(' ', font, font_size)  # учёт пробела
        else:
            lines.append(' '.join(current_line))
            current_line = [word]
            current_width = word_width  # новый размер строки с текущим словом

    # Добавляем последнюю строку, если она не пустая
    if current_line:
        lines.append(' '.join(current_line))

    return lines

